import pytest
from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
import openpyxl
from selenium.webdriver.common.keys import Keys


def test_openpyxl_app():
    op = openpyxl.load_workbook("data_set.xlsx")
    data = op.active
    driver = webdriver.Chrome(executable_path='./chromedriver.exe')
    driver.get('https://bilet.railways.kz/login')
    driver.maximize_window()
    driver.find_element(By.ID, "username").send_keys('vip.nurasyl7@gmail.com')
    driver.find_element(By.ID, "password").send_keys('Nura123!')
    driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div/div/div[2]/div[1]/form/button").click()
    from_city = data.cell(row=2, column=1).value
    fromfill = driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div[2]/div/div/div/div[3]/form/div[1]/div[1]/input[2]")
    fromfill.send_keys(from_city)
    sleep(2)
    fromfill.send_keys(Keys.ENTER)
    to_city = data.cell(row=2, column=2).value
    tofill = driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div[2]/div/div/div/div[3]/form/div[2]/div[1]/input[2]")
    tofill.send_keys(to_city)
    sleep(2)
    fromfill.send_keys(Keys.ENTER)
    date_from = data.cell(row=2, column=3).value
    date = driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div[2]/div/div/div/div[3]/form/div[3]/div[1]/div[1]/div/input')
    date.clear()
    date.send_keys(date_from)
    sleep(1)
    date.send_keys(Keys.ENTER)
    driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div[2]/div/div/div/div[3]/form/button").click()
    driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div[4]/table/tbody/tr[1]/td[5]/table/tfoot/tr/td/button").click()
    driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div[4]/div/table/tbody/tr[1]").click()
    driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div[4]/div/table/tbody/tr[2]/td/div[2]/div/div[1]/div[2]/div/div[3]/div[4]").click()
    driver.find_element(By.XPATH, "/html/body/div[2]/div[3]/div[4]/div/table/tbody/tr[2]/td/button").click()
    sleep(1)

    driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/form/div[4]/div/div/div[3]/div[3]/input').send_keys("0"+str(data.cell(row=2, column=4).value))#D
    driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/form/div[4]/div/div/div[4]/div[1]/input').send_keys(data.cell(row=2, column=5).value)#S
    driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/form/div[4]/div/div/div[4]/div[2]/input').send_keys(data.cell(row=2, column=6).value)#N
    driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/form/div[4]/div/div/div[4]/div[3]/input').send_keys(data.cell(row=2, column=7).value)#F
    driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/form/div[4]/div/div/div[5]/div[1]/input').send_keys(data.cell(row=2, column=8).value)#iin
    driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/form/div[4]/div/div/div[6]/div[1]/input').send_keys(data.cell(row=2, column=9).value)
    driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/form/div[5]/div[2]/div/div/label').click()
    driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/form/div[5]/div[3]/div/button').click()